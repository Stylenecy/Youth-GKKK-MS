#!/usr/bin/env python3
"""
Import anggota pemuda dari DATA_PEMUDA-GKKK-YK.xlsx ke format SQL INSERT.

Aturan privasi (sesuai BRIEF-AI_YGMS.md section1 & section4 dan
BRIEF-AI_YGMS-Kontak-WhatsApp.md section2):
- No. telepon DINORMALISASI ke format 62... lalu disimpan (T1 brief kontak);
  nomor yang gagal dinormalisasi di-kosongkan dan DICATAT JUMLAHNYA SAJA
  di laporan -- nomornya sendiri TIDAK pernah masuk laporan.
- Output berisi nomor asli -> WAJIB masuk folder scripts/import/output/
  yang di-gitignore. Jangan pernah commit berkas output impor.
- Tanggal lahir TIDAK boleh masuk seed yang ter-commit.
- Nama asli TIDAK boleh masuk seed -- hanya nickname yang di-authorized.

Cara pakai:
  python scripts/import/import_members.py --mode sql    -> cetak SQL INSERT
  python scripts/import/import_members.py --mode report -> cetak ringkasan mapping
  python scripts/import/import_members.py --mode json   -> cetak JSON untuk Supabase API
  python scripts/import/import_members.py --mode sql --output scripts/import/output/import_members.sql

Output ke stdout, supaya bisa diarahkan ke file atau pipe ke psql.
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from id_map import pid

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl belum terinstall: pip install openpyxl")

DB_DIR = Path(__file__).resolve().parent.parent.parent / "database"
XLSX_PATH = DB_DIR / "DATA_PEMUDA-GKKK-YK.xlsx"
SHEET_NAME = "ANGGOTA PEMUDA"
# The roster's status column ("Keterangan") lives here, not on ANGGOTA PEMUDA.
ABSENSI_SHEET = "ABSENSI 2026 (Agustus - Desembe"
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent / "output"

SKILL_CATEGORIES = {
    "WL": "worship_leader",
    "SINGER": "singer",
    "PEMUSIK": "musician",
    "SOUND": "sound",
    "MULTIMEDIA": "multimedia",
    "USHER": "usher",
}


def normalize_phone(raw):
    """Normalisasi nomor ke format 62... (aturan T1 brief kontak).

    Buang semua non-digit -> 0 di depan jadi 62 -> 62 di depan dibiarkan ->
    +62 jadi 62. Hasil <10 atau >15 digit dianggap bukan nomor -> None.
    Hasil yang tidak berawalan 628 (bukan nomor seluler Indonesia -- mis.
    nomor rumah 6221.../6222...) juga -> None, karena tombol WhatsApp cuma
    berguna untuk nomor seluler.
    Versi mentah sengaja TIDAK disimpan: brief melarang menambah kolom
    yang tidak diminta, dan nilai mentah tak berguna setelah normalisasi.
    """
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None
    if digits.startswith("0"):
        digits = digits[1:]
    if not digits.startswith("62"):
        digits = "62" + digits
    if not digits.startswith("628"):
        return None
    if len(digits) < 10 or len(digits) > 15:
        return None
    return digits


def parse_date(raw):
    """Parse berbagai format tanggal ke ISO date string atau None."""
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d %B %Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    match = re.match(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s)
    if match:
        months = {
            "januari": 1, "februari": 2, "maret": 3, "april": 4,
            "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
            "september": 9, "oktober": 10, "november": 11, "desember": 12,
        }
        day, mon_str, year = match.groups()
        mon = months.get(mon_str.lower())
        if mon:
            return f"{year}-{mon:02d}-{int(day):02d}"
    return None


def parse_skills(raw):
    """Parse bidang pelayanan menjadi list skill category."""
    if not raw:
        return []
    text = str(raw).upper()
    found = []
    for key, value in SKILL_CATEGORIES.items():
        if key in text:
            found.append(value)
    return found


def name_key(raw):
    """Kunci pencocokan nama antar-sheet.

    Kedua sheet menulis nama orang yang sama dengan cara berbeda: 'Arion
    Sudibyo' vs 'Arion Sudibyo (Arion)', 'Ding-ding' vs 'Ding-Ding'. Mencocokkan
    string mentah hanya berhasil untuk 52 dari 93 orang. Kunci ini membuang
    bagian dalam kurung, merapatkan spasi, dan menyamakan huruf besar-kecil.
    """
    if not raw:
        return ""
    text = re.sub(r"\(.*?\)", " ", str(raw))
    return re.sub(r"\s+", " ", text).strip().lower()


def load_keterangan(wb):
    """Petakan nama -> teks 'Keterangan' dari sheet ABSENSI terbaru.

    Kolom ini TIDAK ada di sheet ANGGOTA PEMUDA -- hanya di sheet absensi --
    sehingga versi lama importer ini tidak pernah melihatnya dan menandai
    seluruh 93 anggota sebagai 'active'. Isinya alasan seseorang sedang tidak
    hadir rutin: 'Kuliah di luar', 'Pindah gereja', 'Ke Belanda', 'Skripsi'.

    Nama yang setelah dinormalisasi cocok ke lebih dari satu baris dibuang,
    bukan ditebak. Proyek ini sudah pernah tertipu pencarian nama yang ambigu
    (dua 'Nathan', dua 'Dian'); menandai orang yang salah sebagai 'pindah
    gereja' jauh lebih buruk daripada tidak menandai siapa pun.
    """
    if ABSENSI_SHEET not in wb.sheetnames:
        return {}, []
    ws = wb[ABSENSI_SHEET]
    seen = {}
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        num, full_name, _nick, ket = row[0], row[1], row[2], row[3]
        if not isinstance(num, (int, float)) or not full_name:
            continue
        key = name_key(full_name)
        if not key:
            continue
        seen.setdefault(key, []).append(str(ket).strip() if ket else "")

    ambiguous = sorted(k for k, v in seen.items() if len(v) > 1)
    out = {k: v[0] for k, v in seen.items() if len(v) == 1}
    return out, ambiguous


def unmatched_report(members, keterangan):
    """Nama di ANGGOTA PEMUDA yang tidak punya pasangan persis di ABSENSI.

    Sengaja TIDAK dicocokkan secara fuzzy. Kandidat terdekat untuk 'christian
    natanael' adalah 'christian yang' -- dua orang berbeda yang lolos ambang
    kemiripan 0,75. Menebak di sini berarti menandai orang yang salah sebagai
    pindah gereja. Yang tidak cocok tetap 'active' (default paling aman), dan
    namanya dilaporkan supaya ejaan di Excel bisa dirapikan di sumbernya.
    """
    return [m["full_name"] for m in members if name_key(m["full_name"]) not in keterangan]


def status_from_keterangan(text):
    """Terjemahkan teks Keterangan jadi (status, catatan).

    Sengaja konservatif: satu-satunya hal yang menurunkan seseorang jadi
    'inactive' adalah keterangan bahwa dia sudah PINDAH GEREJA. Semua alasan
    lain (kuliah di luar, kerja, skripsi, jarang datang) berarti 'away' --
    masih anggota Pemuda, hanya sedang tidak hadir rutin. Salah menandai
    seseorang 'tidak aktif' padahal dia masih jemaat jauh lebih merugikan
    daripada menandainya 'berhalangan'.
    """
    if not text or text == "-":
        return "active", None
    lowered = text.lower()
    if "pindah gereja" in lowered or lowered.strip() == "pindah":
        return "inactive", text
    return "away", text


def load_members(path=XLSX_PATH):
    """Baca sheet ANGGOTA PEMUDA, return list of dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]
    keterangan, ambiguous_names = load_keterangan(wb)
    members = []
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        num = row[0]
        if num is None or not isinstance(num, (int, float)):
            continue
        full_name = str(row[1]).strip() if row[1] else ""
        nickname = str(row[2]).strip() if row[2] else full_name
        # row[3] = No. Telepon -- dinormalisasi (T1 brief kontak WhatsApp);
        # yang gagal di-kosongkan dan dihitung di laporan, bukan disalin.
        raw_phone = row[3]
        whatsapp = normalize_phone(raw_phone)
        has_raw = raw_phone is not None and str(raw_phone).strip() != ""
        hometown = str(row[4]).strip() if row[4] else None
        birth_date = parse_date(row[5])
        skills = parse_skills(row[6])
        if not full_name:
            continue
        status, notes = status_from_keterangan(keterangan.get(name_key(full_name)))
        members.append({
            "num": int(num),
            "full_name": full_name,
            "nickname": nickname,
            "whatsapp": whatsapp,
            "has_raw_phone": has_raw,
            "hometown": hometown,
            "birth_date": birth_date,
            "skills": skills,
            "status": status,
            "notes": notes,
        })
    return members


def generate_sql(members):
    """Generate SQL INSERT statements untuk tabel profiles & skills."""
    lines = [
        "-- Generated: " + datetime.now().isoformat(),
        "-- Sumber: DATA_PEMUDA-GKKK-YK.xlsx -> ANGGOTA PEMUDA",
        "-- Jumlah: " + str(len(members)) + " anggota",
        "--",
        "-- CATATAN PRIVASI:",
        "--   - No. telepon sudah DINORMALISASI ke format 62... (T1 brief kontak).",
        "--   - Tanggal lahir ada di sini untuk keperluan internal,",
        "--     TIDAK boleh masuk seed.ts yang ter-commit.",
        "--   - Berkas ini berisi nomor asli: JANGAN commit.",
        "--     Jalankan dengan --output scripts/import/output/ (di-gitignore).",
        "--",
        "-- URUTAN: import_members.sql -> import_cross.sql -> import_finance.sql -> import_events.sql",
        "--",
        "BEGIN;",
        "",
        "-- 1. Profiles (id pakai nomor urut sebagai pseudo-UUID yang deterministik)",
        "--    INSERT ini idempotent: aman diulang.",
    ]
    for m in members:
        fn = m["full_name"].replace("'", "''")
        nn = m["nickname"].replace("'", "''")
        ht = m["hometown"].replace("'", "''") if m["hometown"] else None
        bd = m["birth_date"] or "NULL"
        wa = m["whatsapp"]
        ht_sql = f"'{ht}'" if ht else "NULL"
        bd_sql = f"'{bd}'" if bd != "NULL" else "NULL"
        wa_sql = f"'{wa}'" if wa else "NULL"
        status = m["status"]
        notes_sql = f"'{m['notes'].replace(chr(39), chr(39) * 2)}'" if m["notes"] else "NULL"
        row_id = pid(f"p{m['num']:03d}")
        lines.append(
            f"INSERT INTO public.profiles (id, full_name, nickname, birth_date, hometown, whatsapp, status, notes, is_active) "
            f"VALUES ('{row_id}', '{fn}', '{nn}', {bd_sql}, {ht_sql}, {wa_sql}, '{status}', {notes_sql}, true) "
            f"ON CONFLICT (id) DO UPDATE SET "
            f"full_name = EXCLUDED.full_name, nickname = EXCLUDED.nickname, "
            f"birth_date = EXCLUDED.birth_date, hometown = EXCLUDED.hometown, "
            f"whatsapp = EXCLUDED.whatsapp, notes = EXCLUDED.notes, "
            f"status = EXCLUDED.status, updated_at = now();"
        )
    lines.append("")
    lines.append("-- 2. Skills (hapus dulu yang lama, insert yang baru -- idempotent)")
    for m in members:
        row_id = pid(f"p{m['num']:03d}")
        lines.append(
            f"DELETE FROM public.skills WHERE profile_id = '{row_id}';"
        )
        for skill in m["skills"]:
            lines.append(
                f"INSERT INTO public.skills (profile_id, category, skill_name, proficiency_level) "
                f"VALUES ('{row_id}', '{skill}', '{skill}', 'intermediate');"
            )
    lines.append("")
    lines.append("COMMIT;")
    return "\n".join(lines)


def generate_json(members):
    """Generate JSON array untuk Supabase API."""
    output = []
    for m in members:
        fn = m["full_name"]
        nn = m["nickname"]
        ht = m["hometown"]
        bd = m["birth_date"]
        output.append({
            "id": pid(f"p{m['num']:03d}"),
            "full_name": fn,
            "nickname": nn,
            "birth_date": bd,
            "hometown": ht,
            "whatsapp": m["whatsapp"],
            "status": m["status"],
            "notes": m["notes"],
            "is_active": True,
        })
    return json.dumps(output, ensure_ascii=False, indent=2)


def generate_report(members):
    """Print ringkasan data untuk review Dex."""
    with_phone = sum(1 for m in members if m["whatsapp"])
    # Jumlah saja. Nomor yang gagal TIDAK boleh dicetak di laporan (brief).
    failed_raw = sum(1 for m in members if m["whatsapp"] is None and m["has_raw_phone"])
    print(f"Total anggota: {len(members)}")
    print(f"Dengan nomor WhatsApp tersimpan (ternormalisasi): {with_phone}")
    print(f"Dengan tanggal lahir: {sum(1 for m in members if m['birth_date'])}")
    print(f"Dengan skills: {sum(1 for m in members if m['skills'])}")
    print(f"Dengan hometown: {sum(1 for m in members if m['hometown'])}")
    print(f"Nomor GAGAL dinormalisasi (dikosongkan): {failed_raw}")
    print()
    counts = {}
    for m in members:
        counts[m["status"]] = counts.get(m["status"], 0) + 1
    print("Status (dari kolom Keterangan sheet ABSENSI):")
    for key in ("active", "away", "inactive", "alumni"):
        if key in counts:
            print(f"  {key}: {counts[key]}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    keterangan, ambiguous = load_keterangan(wb)
    missing = unmatched_report(members, keterangan)
    print(f"  (tanpa pasangan di sheet ABSENSI -> dianggap active: {len(missing)})")
    if ambiguous:
        print(f"  NAMA AMBIGU di ABSENSI, sengaja dilewati: {', '.join(ambiguous)}")
    if missing:
        print()
        print("Belum cocok ke sheet ABSENSI -- rapikan ejaannya di Excel supaya")
        print("keterangannya ikut terbaca (TIDAK dicocokkan otomatis, terlalu berisiko):")
        for name in missing:
            print(f"  - {name}")
    print()
    print("Sample (5 pertama):")
    for m in members[:5]:
        print(f"  {m['num']:3d}. {m['full_name'][:40]:40s} | {m['nickname']:15s} | {m['birth_date'] or 'N/A':10s} | wa: {'OK' if m['whatsapp'] else '-'} | skills: {', '.join(m['skills']) or '-'}")
    print()
    print("Nama duplikat (jika ada):")
    names = [m["nickname"] for m in members]
    seen = set()
    dupes = []
    for n in names:
        if n in seen:
            dupes.append(n)
        seen.add(n)
    if dupes:
        for d in dupes:
            print(f"  - {d}")
    else:
        print("  (tidak ada)")


def has_raw_phone(member):
    """True bila baris sumber berisi nomor namun hasil normalisasi None."""
    return False  # diganti nilai asli saat load: lihat load_members()



def main():
    parser = argparse.ArgumentParser(description="Import anggota pemuda dari Excel")
    parser.add_argument("--mode", choices=["sql", "json", "report"], default="report")
    parser.add_argument(
        "--output",
        help="Output file path (default: scripts/import/output/ -- DI-GITIGNORE)",
    )
    args = parser.parse_args()

    members = load_members()
    if args.mode == "sql":
        output = generate_sql(members)
    elif args.mode == "json":
        output = generate_json(members)
    else:
        generate_report(members)
        return

    if args.output:
        out_path = Path(args.output)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(output, encoding="utf-8")
        print(f"Output tertulis ke: {out_path}", file=sys.stderr)
    else:
        # stdout biar bisa dipipe ke psql. Kalau mau menyimpan ke berkas,
        # arahkan ke scripts/import/output/ yang DI-GITIGNORE supaya nomor
        # asli tidak pernah ter-commit.
        print(output)


if __name__ == "__main__":
    main()
