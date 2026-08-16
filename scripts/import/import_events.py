#!/usr/bin/env python3
"""
Import jadwal ibadah & penatalayan dari Jadwal Penatalayan Pemuda_.xlsx (tab 2026).

Mapping kolom:
  - Minggu 1-5 per bulan
  - Tanggal -> date
  - Worship Leader -> steward role "WL"
  - Singer -> steward role "Singer"
  - Pemusik -> steward role "Pemusik"
  - Multimedia -> steward role "Multimedia"
  - Sound -> steward role "Sound"
  - Usher -> steward role "Usher"
  - Tema -> weekly_theme

Cara pakai:
  python scripts/import/import_events.py --mode sql    -> cetak SQL INSERT
  python scripts/import/import_events.py --mode report -> cetak ringkasan
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl belum terinstall: pip install openpyxl")

DB_DIR = Path(__file__).resolve().parent.parent.parent / "database"
XLSX_PATH = DB_DIR / "Jadwal Penatalayan Pemuda_.xlsx"
SHEET_NAME = "2026"

ROLE_COLUMNS = {
    3: "WL",
    4: "Singer",
    5: "Pemusik",
    6: "Multimedia",
    7: "Sound",
    8: "Usher",
}


def load_schedule(path=XLSX_PATH):
    """Baca jadwal 2026, return list of dict dengan date, theme, stewards."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET_NAME]
    events = []
    current_month = None
    month_names = {
        "januari": 1, "februari": 2, "maret": 3, "april": 4,
        "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12,
    }
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if row[1] is None and row[2] is None:
            continue
        cell_b = str(row[1]).strip().lower() if row[1] else ""
        if cell_b in month_names:
            current_month = month_names[cell_b]
            continue
        if cell_b == "minggu" or (row[2] is None and not isinstance(row[1], (int, float))):
            continue
        date_val = row[2]
        if date_val is None or not isinstance(date_val, datetime):
            continue
        if current_month is None:
            continue
        date_str = date_val.strftime("%Y-%m-%d")
        week_num = int(row[1]) if isinstance(row[1], (int, float)) else None
        theme = str(row[9]).strip() if row[9] else None
        stewards = {}
        for col_idx, role in ROLE_COLUMNS.items():
            names_raw = row[col_idx]
            if names_raw:
                raw_str = str(names_raw).strip()
                if raw_str and raw_str.lower() not in ("none", "-", ""):
                    names = [n.strip() for n in raw_str.replace(",", "\n").split("\n") if n.strip() and n.strip() != "-"]
                    if names:
                        stewards[role] = names
        events.append({
            "date": date_str,
            "month": current_month,
            "week": week_num,
            "theme": theme,
            "stewards": stewards,
        })
    return events


def generate_sql(events):
    """Generate SQL INSERT untuk events & steward_assignments."""
    lines = [
        "-- Generated: " + datetime.now().isoformat(),
        "-- Sumber: Jadwal Penatalayan Pemuda_.xlsx -> tab 2026",
        "-- Jumlah: " + str(len(events)) + " ibadah",
        "--",
        "BEGIN;",
        "",
        "-- 1. Events",
    ]
    for i, ev in enumerate(events, 1):
        theme = (ev["theme"] or "Ibadah Pemuda").replace("'", "''")
        event_type = "cross" if ev["theme"] and "cross" in ev["theme"].lower() else "worship"
        lines.append(
            f"INSERT INTO public.events (id, date, weekly_theme, event_type, status, description) "
            f"VALUES ('e{i:03d}', '{ev['date']}T17:00:00Z', '{theme}', '{event_type}', 'published', "
            f"'Minggu {ev['week']} bulan {ev['month']}') "
            f"ON CONFLICT (id) DO UPDATE SET "
            f"date = EXCLUDED.date, weekly_theme = EXCLUDED.weekly_theme, updated_at = now();"
        )
    lines.append("")
    lines.append("-- 2. Steward assignments")
    s_id = 0
    for i, ev in enumerate(events, 1):
        for role, names in ev["stewards"].items():
            for name in names:
                s_id += 1
                lines.append(
                    f"INSERT INTO public.steward_assignments (id, event_id, profile_id, role, status) "
                    f"SELECT 's{s_id:04d}', 'e{i:03d}', id, '{role}', 'assigned' "
                    f"FROM public.profiles WHERE nickname = '{name}' "
                    f"ON CONFLICT DO NOTHING;"
                )
    lines.append("")
    lines.append("COMMIT;")
    return "\n".join(lines)


def generate_report(events):
    """Print ringkasan jadwal."""
    print(f"Total ibadah terjadwal: {len(events)}")
    months = {}
    for ev in events:
        months[ev["month"]] = months.get(ev["month"], 0) + 1
    print(f"\nPer bulan:")
    month_labels = {1:"Jan", 2:"Feb", 3:"Mar", 4:"Apr", 5:"Mei", 6:"Jun",
                    7:"Jul", 8:"Agu", 9:"Sep", 10:"Okt", 11:"Nov", 12:"Des"}
    for m in sorted(months.keys()):
        print(f"  {month_labels.get(m, '?'):>3s}: {months[m]} ibadah")
    print(f"\nSample (5 pertama):")
    for ev in events[:5]:
        n_stewards = sum(len(v) for v in ev["stewards"].values())
        print(f"  {ev['date']}  {ev['theme'][:40] if ev['theme'] else '(tanpa tema)':40s}  [{n_stewards} penatalayan]")


def main():
    parser = argparse.ArgumentParser(description="Import jadwal ibadah & penatalayan")
    parser.add_argument("--mode", choices=["sql", "report"], default="report")
    parser.add_argument("--output", help="Output file path (default: stdout)")
    args = parser.parse_args()

    events = load_schedule()
    if args.mode == "sql":
        output = generate_sql(events)
        if args.output:
            Path(args.output).write_text(output, encoding="utf-8")
            print(f"Output tertulis ke: {args.output}", file=sys.stderr)
        else:
            print(output)
    else:
        generate_report(events)


if __name__ == "__main__":
    main()
